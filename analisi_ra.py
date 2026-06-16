import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil, glob, os, re

INPUT_DIR = "input/"
OUTPUT_DIR = "output/"
SHEET_NAME = "Estadístiques alumnes (RA)"
STATS_SHEET_NAME = "Estadístiques grup (MP)"

os.makedirs(OUTPUT_DIR, exist_ok=True)
FILES = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))

def parse_filename(filepath):
    m = re.search(r'_av_(\d+)_\d{4}-\d{2}-\d{2}_(.+)\.xlsx$', os.path.basename(filepath))
    if m:
        return int(m.group(1)), m.group(2)
    return None, None

def merge_av_data(df_prev, df_curr):
    id_col = df_curr.columns[0]
    nom_col = df_curr.columns[1]
    ra_prev = {c for c in df_prev.columns if str(c).endswith("RA")}
    ra_curr = {c for c in df_curr.columns if str(c).endswith("RA")}
    common_ra = ra_prev & ra_curr

    prev_by_id = df_prev.set_index(df_prev.columns[0])

    for idx, row in df_curr.iterrows():
        student_id = row[id_col]
        if student_id not in prev_by_id.index:
            continue
        prev_row = prev_by_id.loc[student_id]
        for col in common_ra:
            if not is_evaluated(row[col]) and is_evaluated(prev_row[col]):
                val = prev_row[col]
                if isinstance(val, str) and df_curr[col].dtype != object:
                    df_curr[col] = df_curr[col].astype(object)
                df_curr.at[idx, col] = val

    curr_ids = set(df_curr[id_col])
    missing = df_prev[~df_prev[df_prev.columns[0]].isin(curr_ids)]
    if not missing.empty:
        new_rows = []
        for _, student in missing.iterrows():
            new_row = {col: pd.NA for col in df_curr.columns}
            new_row[id_col] = student[df_prev.columns[0]]
            new_row[nom_col] = student[df_prev.columns[1]]
            for col in common_ra:
                if is_evaluated(student[col]):
                    new_row[col] = student[col]
            new_rows.append(new_row)
        df_curr = pd.concat([df_curr, pd.DataFrame(new_rows, columns=df_curr.columns)], ignore_index=True)

    return df_curr

def is_evaluated(val):
    """Avaluat = número o 'NA' (suspès). PDT i buit = no avaluat."""
    if pd.isna(val):
        return False
    if isinstance(val, str):
        return val.strip().upper() == "NA"
    return True

def is_approved(val):
    """Aprovat = nota >= 5. 'NA' = avaluat però suspès."""
    if pd.isna(val):
        return False
    if isinstance(val, str):
        return False
    try:
        return float(val) >= 5
    except (ValueError, TypeError):
        return False

def get_mp_stats(df, ra_cols):
    mp_codes = sorted(set(c.split("_")[0] for c in ra_cols))
    mp_ra = {mp: [c for c in ra_cols if c.split("_")[0] == mp] for mp in mp_codes}
    stats = {mp: {"aproven": 0, "suspenen": 0} for mp in mp_codes}
    for _, row in df.iterrows():
        for mp, cols in mp_ra.items():
            evaluated = [c for c in cols if is_evaluated(row[c])]
            if not evaluated:
                continue
            if all(is_approved(row[c]) for c in evaluated):
                stats[mp]["aproven"] += 1
            else:
                stats[mp]["suspenen"] += 1
    return stats


def write_stats_sheet(wb, mp_stats):
    if STATS_SHEET_NAME in wb.sheetnames:
        del wb[STATS_SHEET_NAME]
    ws = wb.create_sheet(STATS_SHEET_NAME)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    label_font = Font(name="Arial", bold=True)
    data_font = Font(name="Arial")
    header_fill = PatternFill("solid", start_color="2E4057")
    suspenen_fill = PatternFill("solid", start_color="F4CCCC")
    aproven_fill = PatternFill("solid", start_color="D9EAD3")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    mp_codes = sorted(mp_stats.keys())

    # Row 1: headers
    ws.cell(row=1, column=1, value="").border = border
    ws.column_dimensions["A"].width = 12
    for col_idx, mp in enumerate(mp_codes, start=2):
        cell = ws.cell(row=1, column=col_idx, value=mp)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    ws.row_dimensions[1].height = 22

    # Row 2: Suspenen
    cell = ws.cell(row=2, column=1, value="Suspenen")
    cell.font = label_font
    cell.fill = suspenen_fill
    cell.alignment = center
    cell.border = border
    for col_idx, mp in enumerate(mp_codes, start=2):
        cell = ws.cell(row=2, column=col_idx, value=mp_stats[mp]["suspenen"])
        cell.font = data_font
        cell.fill = suspenen_fill
        cell.alignment = center
        cell.border = border

    # Row 3: Aproven
    cell = ws.cell(row=3, column=1, value="Aproven")
    cell.font = label_font
    cell.fill = aproven_fill
    cell.alignment = center
    cell.border = border
    for col_idx, mp in enumerate(mp_codes, start=2):
        cell = ws.cell(row=3, column=col_idx, value=mp_stats[mp]["aproven"])
        cell.font = data_font
        cell.fill = aproven_fill
        cell.alignment = center
        cell.border = border


def process_file(filepath, df=None):
    if df is None:
        df = pd.read_excel(filepath, header=1, keep_default_na=False, na_values=[""])
    ra_cols = [c for c in df.columns if str(c).endswith("RA")]

    results = []
    for _, row in df.iterrows():
        student_id = row.iloc[0]
        student_name = row.iloc[1]
        evaluats = sum(is_evaluated(row[c]) for c in ra_cols)
        aprovats = sum(is_approved(row[c]) for c in ra_cols)
        results.append({
            "ID Alumne": student_id,
            "Nom": student_name,            
            "RA Aprovats": aprovats,
            "RA Avaluats": evaluats,
        })

    filename = os.path.basename(filepath)
    outpath = os.path.join(OUTPUT_DIR, filename)
    shutil.copy(filepath, outpath)

    wb = load_workbook(outpath)
    mp_stats = get_mp_stats(df, ra_cols)
    write_stats_sheet(wb, mp_stats)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E4057")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="EBF0F7")

    headers = ["ID Alumne", "Nom", "RA Aprovats", "RA Avaluats"]
    col_widths = [12, 30, 16, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    for row_idx, record in enumerate(results, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record[key])
            cell.border = border
            cell.alignment = center if col_idx != 2 else Alignment(vertical="center")
            if fill:
                cell.fill = fill
            cell.font = Font(name="Arial")

    wb.save(outpath)

print("""
EsferaPowerToys-Analisi-RA
v1.2.0
""")

if not FILES:
    print(f"No s'han trobat fitxers .xlsx a '{INPUT_DIR}'")
else:
    av2_map = {}
    for f in FILES:
        av_num, class_name = parse_filename(f)
        if av_num == 2 and class_name:
            av2_map[class_name] = f

    for f in sorted(FILES):
        av_num, class_name = parse_filename(f)
        if av_num == 3 and class_name in av2_map:
            df_prev = pd.read_excel(av2_map[class_name], header=1, keep_default_na=False, na_values=[""])
            df_curr = pd.read_excel(f, header=1, keep_default_na=False, na_values=[""])
            df_merged = merge_av_data(df_prev, df_curr)
            process_file(f, df=df_merged)
        else:
            process_file(f)
    print(f"Fitxers processats. Podeu trobar els resultats a la carpeta '{OUTPUT_DIR}' (s'han creat pestanyes noves amb les dades demanades).")
